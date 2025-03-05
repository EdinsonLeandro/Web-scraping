import pandas as pd
import os, gspread, os
from scrapy.utils.project import get_project_settings
from scrapy.crawler import CrawlerProcess
from google.oauth2 import service_account
from gspread.exceptions import APIError
from gspread_dataframe import set_with_dataframe
from openpyxl import load_workbook


def run_spider(spider_name, filename, *args):

    # Get project settings
    project_settings = get_project_settings()

    # Rename filename
    feed_name = project_settings.attributes['FEEDS'].value.attributes
    feed_name[filename] = feed_name.pop('Name.csv')

    # Set settings and start
    process = CrawlerProcess(settings=project_settings)

    # Check if there are additional arguments.
    if len(args[0]) > 0:
        # Tuple unpacking
        args = args[0]

        # Run spider
        process.crawl(spider_name, args)
    else:
        # Run spider
        process.crawl(spider_name)

    process.start()

    # After the process, read the file and drop rows just in case you have lines equals to the title.
    # Read file
    temp = pd.read_csv(filename, encoding='utf-8')

    # Drop rows
    column_name = temp.columns[0]
    indexes_to_drop = temp[temp[column_name] == column_name].index

    if indexes_to_drop:
        temp.drop(indexes_to_drop, inplace=True)

        # Save .csv file
        temp.to_csv(filename, index=False, encoding='utf-8')


def get_credentials(platform):
    """
    Fetch credentials for the specified platform.
    """
    # Location
    path='Data/Credentials/credentials.csv'

    # Open file
    credentials = pd.read_csv(path, dtype=str)

    # Get credentials
    credentials = credentials.loc[credentials['Platform'] == platform, ['Username', 'Password']]
    
    return credentials.values[0]


def prepare_raw_data(filename1, filename2, drop_rows=False):
    """
    Delete, if necessary, rows of raw data in case you need to run a script more than once.
    """

    # Read data
    df = pd.read_csv(filename1)

    if drop_rows:
        # Read and clean statistics file
        temp = pd.read_csv(filename2, encoding='utf-8')

        # Drop rows from the user data to continue script execution
        df.drop(range(temp.shape[0]), axis=0, inplace=True)

    return df


def authorize_google():
    # Find the path of Google Sheets key
    path = r'/PrivateKeys/gstats_google_sheet_key.json'

    try:
        service_file = os.path.normpath(os.getcwd() + (os.sep + os.pardir) * 1) + path

        # Set scopes and credentials
        scopes = ['https://www.googleapis.com/auth/spreadsheets']
        creds = service_account.Credentials.from_service_account_file(service_file, scopes=scopes)

    except FileNotFoundError as error_msg:
        # I am using this function from different paths.
        service_file = os.path.normpath(os.getcwd() + (os.sep + os.pardir) * 2) + path

        # Set scopes and credentials
        scopes = ['https://www.googleapis.com/auth/spreadsheets']
        creds = service_account.Credentials.from_service_account_file(service_file, scopes=scopes)

    return gspread.authorize(creds)     #type: ignore


def send_output_file(input_file, output_file, tab_name):
    # Open input file
    df = pd.read_csv(input_file)

    if "docs.google.com" in output_file:
        # Authorize
        gc = authorize_google()

        print('Sending data to Google Sheet.\n')

        # Open url
        sheet_file = gc.open_by_url(output_file)

        try:
            sheet_file.add_worksheet(title=tab_name, rows="50", cols="5")   #type: ignore
            
            # Send info to empty sheet
            set_with_dataframe(sheet_file.worksheet(tab_name), df)

        except APIError as error_msg:
            # The sheet already exist
            print('The sheet already exist: {}'.format(error_msg))

            # Replane NaN values with empty string. Otherwise "append_row" returns error
            df.fillna("", inplace=True)
                    
            # https://stackoverflow.com/questions/45540827/appending-pandas-data-frame-to-google-spreadsheet
            # Append all rows to Google Sheet file.
            sheet_file.worksheet(tab_name).append_rows(df.values.tolist())

    else:
        if not os.path.isfile(output_file):
            df.to_excel(output_file, sheet_name=tab_name, startrow= 0, index=False)

        else:
            # Load the Excel file into memory.
            wb = load_workbook(output_file)
            writer = pd.ExcelWriter(output_file, engine='openpyxl')

            if tab_name in wb.sheetnames:
                startrow = wb[tab_name].max_row

            df.to_excel(writer, sheet_name=tab_name, startrow= startrow, index=False)

            wb.save(output_file)

    print('Data was succesfully saved to a file.')
